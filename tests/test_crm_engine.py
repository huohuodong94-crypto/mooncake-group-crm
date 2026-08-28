#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""门店团购 CRM 引擎单测（stdlib unittest + importlib 加载脚本纯函数）。

运行：python3 -m unittest tests/test_crm_engine.py   （在 skill 根目录）
"""
import importlib.util
import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

SKILL_DIR = Path(__file__).resolve().parent.parent
SCRIPTS = SKILL_DIR / "scripts"
TESTS = SKILL_DIR / "tests"
TODAY = "2026-07-31"


def _load(modname, filename):
    spec = importlib.util.spec_from_file_location(modname, SCRIPTS / filename)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


crm = _load("crm_under_test", "crm.py")
validate = _load("validate_under_test", "validate_mooncake-group-crm_output.py")
RULES = crm.load_rules()


def fresh_ledger():
    return crm.empty_ledger()


def args_for(argv):
    return crm.build_parser().parse_args(argv)


class TestRulesContract(unittest.TestCase):
    def test_stages_and_levels_match_template(self):
        self.assertEqual(RULES["intent_levels"], ["A", "B", "C", "D"])
        self.assertEqual(
            RULES["status_machine"]["states"],
            ["初次接触", "需求确认", "方案报价", "谈判中", "已签约", "已回款", "已流失"],
        )

    def test_export_columns_count_matches_template(self):
        self.assertEqual(len(RULES["export"]["门店客户商机表"]["columns"]), 22)  # V1.2
        self.assertEqual(len(RULES["export"]["总部客户档案表"]["columns"]), 27)

    def test_v12_enums(self):
        self.assertEqual(
            RULES["stores"],
            ["月湖店", "体验中心店", "国金店", "五一店", "花样汇店", "市场部"],
        )
        self.assertEqual(RULES["lead_sources"], ["新获客", "老客转介绍", "老客续费"])
        self.assertEqual(RULES["demand_categories"],
                         ["散装", "68礼盒", "98礼盒", "138礼盒", "定制礼盒", "待定"])
        self.assertEqual(RULES["status_machine"]["closing_states"], ["已签约", "已回款", "已流失"])


class TestReminderAndTransition(unittest.TestCase):
    def test_quote_stage_uses_2_days(self):
        self.assertEqual(
            crm.compute_next_follow_date(RULES, "方案报价", "A", crm.parse_date(TODAY)),
            "2026-08-02",
        )

    def test_intent_level_days(self):
        self.assertEqual(crm.compute_next_follow_date(RULES, "谈判中", "A", crm.parse_date(TODAY)), "2026-08-03")
        self.assertEqual(crm.compute_next_follow_date(RULES, "谈判中", "B", crm.parse_date(TODAY)), "2026-08-05")
        self.assertEqual(crm.compute_next_follow_date(RULES, "谈判中", "C", crm.parse_date(TODAY)), "2026-08-07")

    def test_D_level_no_auto_reminder(self):
        self.assertIsNone(crm.compute_next_follow_date(RULES, "谈判中", "D", crm.parse_date(TODAY)))

    def test_overdue_retry_1_day(self):
        self.assertEqual(
            crm.compute_next_follow_date(RULES, "谈判中", "A", crm.parse_date(TODAY), overdue=True),
            "2026-08-01",
        )

    def test_transitions(self):
        self.assertTrue(crm.transition_allowed(RULES, "初次接触", "需求确认"))
        self.assertTrue(crm.transition_allowed(RULES, "谈判中", "已签约"))
        self.assertTrue(crm.transition_allowed(RULES, "方案报价", "已流失"))
        self.assertFalse(crm.transition_allowed(RULES, "初次接触", "已签约"))
        self.assertFalse(crm.transition_allowed(RULES, "已回款", "谈判中"))
        self.assertFalse(crm.transition_allowed(RULES, "已签约", "已流失"))  # 签约后只能→已回款


class TestAtomicLedger(unittest.TestCase):
    def test_save_load_roundtrip_and_no_tmp_leftover(self):
        with tempfile.TemporaryDirectory() as d:
            path = Path(d) / "ledger.json"
            led = fresh_ledger()
            led["customers"].append({"cust_id": "KH-2026-0001", "name": "测试"})
            crm.save_ledger(path, led)
            loaded = crm.load_ledger(path)
            self.assertEqual(loaded["customers"][0]["name"], "测试")
            leftovers = [p for p in Path(d).iterdir() if p.suffix == ".tmp"]
            self.assertEqual(leftovers, [])

    def test_load_missing_returns_empty(self):
        # 只针对不存在的临时路径测试，绝不触碰生产 ledger（旧版本曾 unlink 生产文件，已修正）
        led = crm.load_ledger("/nonexistent/dir/ledger.json")
        self.assertEqual(led["opportunities"], [])
        self.assertEqual(led["customers"], [])

    def test_next_id_format(self):
        led = fresh_ledger()
        self.assertEqual(crm.next_id(led, "MO", "2026"), "MO-2026-0001")
        self.assertEqual(crm.next_id(led, "MO", "2026"), "MO-2026-0002")
        self.assertEqual(crm.next_id(led, "KH", "2026"), "KH-2026-0001")


class TestValidator(unittest.TestCase):
    def setUp(self):
        self.schema = json.loads((SKILL_DIR / "references" / "output-schema.json").read_text(encoding="utf-8"))
        self.valid = json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8"))
        self.invalid = json.loads((TESTS / "fixture_invalid.json").read_text(encoding="utf-8"))

    def test_valid_passes(self):
        self.assertEqual(validate.validate_record(self.valid, self.schema), [])

    def test_invalid_reports_multiple_errors(self):
        errs = validate.validate_record(self.invalid, self.schema)
        joined = "\n".join(errs)
        self.assertTrue(len(errs) >= 3)
        self.assertIn("intent_level", joined)          # 非法枚举 高
        self.assertIn("intent_level", joined)         # 非法枚举
        self.assertIn("additionalProperties", joined)  # 多余字段 amount/level/phone_number

    def test_cli_exit_codes(self):
        vscript = str(SCRIPTS / "validate_mooncake-group-crm_output.py")
        ok = subprocess.run([sys.executable, vscript, str(TESTS / "fixture_valid.json")], capture_output=True, text=True)
        bad = subprocess.run([sys.executable, vscript, str(TESTS / "fixture_invalid.json")], capture_output=True, text=True)
        self.assertEqual(ok.returncode, 0, ok.stdout + ok.stderr)
        self.assertEqual(bad.returncode, 1, bad.stdout + bad.stderr)


class TestEngineAdd(unittest.TestCase):
    def test_add_creates_customer_and_opportunity(self):
        led = fresh_ledger()
        a = args_for(["--mode", "add", "--confirmed", "--name", "深圳八方科技有限公司", "--contact", "王经理",
                      "--contact-phone", "13800000000", "--store", "体验中心店", "--owner", "张店长",
                      "--source", "老客户转介绍", "--need", "员工中秋福利，预算约150元/份",
                      "--customer-address", "深圳市南山区科技园",
                      "--demand-category", "68礼盒", "--quantity", "300", "--amount", "45000",
                      "--intent-level", "A", "--stage", "方案报价", "--next-date", "2026-08-05",
                      "--follow-content", "已发送报价方案，客户内部确认中"])
        res = crm.cmd_add(a, led, RULES, TODAY)
        self.assertTrue(res["_mutated"])
        self.assertEqual(res["customer_id"], "KH-2026-0001")
        self.assertEqual(res["opp_id"], "MO-2026-0001")
        # V1.2.4：严格按 ABCD 提醒，口述时间不覆盖——方案报价 2 天优先 → 08-02
        self.assertEqual(res["next_follow_date"], "2026-08-02")
        opp = crm.find_opp_by_id(led, "MO-2026-0001")
        self.assertEqual(opp["stage"], "方案报价")
        self.assertEqual(opp["latest_follow"], "已发送报价方案，客户内部确认中")
        self.assertEqual(len(led["follow_log"]), 1)

    def test_add_requires_intent_level(self):
        led = fresh_ledger()
        a = args_for(["--mode", "add", "--name", "某客户", "--store", "月湖店"])
        res = crm.cmd_add(a, led, RULES, TODAY)
        self.assertEqual(res["error"], "missing_field")
        self.assertEqual(led["opportunities"], [])

    def test_second_add_same_customer_new_opportunity(self):
        led = fresh_ledger()
        a1 = args_for(["--mode", "add", "--confirmed", "--name", "复购客户", "--contact", "王总", "--contact-phone", "13900001111",
                       "--intent-level", "B", "--store", "月湖店", "--need", "x", "--source", "新获客",
                       "--customer-address", "长沙市岳麓区", "--demand-category", "98礼盒"])
        r1 = crm.cmd_add(a1, led, RULES, TODAY)
        a2 = args_for(["--mode", "add", "--confirmed", "--name", "复购客户", "--contact", "王总", "--contact-phone", "13900001111",
                       "--intent-level", "A", "--store", "月湖店", "--need", "y", "--source", "老客续费",
                       "--customer-address", "长沙市岳麓区", "--demand-category", "散装"])
        r2 = crm.cmd_add(a2, led, RULES, TODAY)
        self.assertEqual(r1["customer_id"], r2["customer_id"])   # 同一客户
        self.assertNotEqual(r1["opp_id"], r2["opp_id"])          # 不同商机
        self.assertFalse(r2["created_customer"])

    def test_phone_conflict_needs_confirmation(self):
        led = fresh_ledger()
        crm.cmd_add(args_for(["--mode", "add", "--confirmed", "--name", "甲", "--contact", "张三", "--contact-phone", "13900000000",
                              "--intent-level", "A", "--store", "月湖店", "--need", "x", "--source", "新获客",
                              "--customer-address", "长沙", "--demand-category", "散装"]), led, RULES, TODAY)
        res = crm.cmd_add(args_for(["--mode", "add", "--confirmed", "--name", "甲", "--contact", "张三", "--contact-phone", "13811112222",
                                    "--intent-level", "A", "--store", "月湖店", "--need", "y", "--source", "新获客",
                                    "--customer-address", "长沙", "--demand-category", "散装"]), led, RULES, TODAY)
        self.assertEqual(res["status"], "needs_confirmation")
        self.assertEqual(res["reason"], "phone_mismatch")  # 过了确认闸口后才轮到去重判定
        self.assertEqual(len(led["customers"]), 1)  # 未重复建档


class TestConfirmGate(unittest.TestCase):
    """关键控制硬闸口：金额/数量/电话/日期未 --confirmed 一律拒写。"""

    ADD_WITH_AMOUNT = ["--mode", "add", "--name", "闸口客户", "--contact", "李四", "--contact-phone", "13800000000",
                       "--store", "月湖店", "--need", "x", "--intent-level", "A", "--source", "新获客",
                       "--customer-address", "长沙", "--demand-category", "散装",
                       "--quantity", "300", "--amount", "45000", "--next-date", "2026-08-05"]

    def test_add_without_confirmed_refuses_and_no_write(self):
        led = fresh_ledger()
        res = crm.cmd_add(args_for(self.ADD_WITH_AMOUNT), led, RULES, TODAY)
        self.assertEqual(res["status"], "needs_confirmation")
        self.assertEqual(res["reason"], "confirm_required")
        self.assertIn("金额", res["fields"])
        self.assertIn("电话", res["fields"])
        self.assertEqual(led["opportunities"], [])   # 未落库
        self.assertEqual(led["customers"], [])

    def test_add_with_confirmed_writes(self):
        led = fresh_ledger()
        res = crm.cmd_add(args_for(self.ADD_WITH_AMOUNT + ["--confirmed"]), led, RULES, TODAY)
        self.assertTrue(res.get("_mutated"))
        self.assertEqual(len(led["opportunities"]), 1)

    def test_update_stage_only_needs_no_confirm(self):
        led = fresh_ledger()
        crm.apply_record(led, RULES, TODAY,
                         json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8")), confirmed=True)
        # 仅改阶段（无金额/数量/电话/日期）不带 --confirmed 也应放行
        res = crm.cmd_update(args_for(["--mode", "update", "--opp-id", "MO-2026-0001", "--stage", "谈判中"]),
                             led, RULES, TODAY)
        self.assertTrue(res.get("_mutated"), res)
        self.assertEqual(crm.find_opp_by_id(led, "MO-2026-0001")["stage"], "谈判中")

    def test_update_amount_without_confirmed_refuses(self):
        led = fresh_ledger()
        crm.apply_record(led, RULES, TODAY,
                         json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8")), confirmed=True)
        res = crm.cmd_update(args_for(["--mode", "update", "--opp-id", "MO-2026-0001", "--amount", "50000"]),
                             led, RULES, TODAY)
        self.assertEqual(res["status"], "needs_confirmation")
        self.assertEqual(crm.find_opp_by_id(led, "MO-2026-0001")["expected_amount"], 45000)  # 未改

    def test_apply_respects_confirmed_flag(self):
        rec = json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8"))
        led = fresh_ledger()
        refused = crm.apply_record(led, RULES, TODAY, rec, confirmed=False)
        self.assertEqual(refused["status"], "needs_confirmation")
        self.assertEqual(led["opportunities"], [])
        ok = crm.apply_record(led, RULES, TODAY, rec, confirmed=True)
        self.assertTrue(ok.get("_mutated"))
        self.assertEqual(len(led["opportunities"]), 1)


class TestApplyAndLifecycle(unittest.TestCase):
    def _add(self, led, record, confirmed=True):
        return crm.apply_record(led, RULES, TODAY, record, confirmed=confirmed)

    def test_full_lifecycle(self):
        led = fresh_ledger()
        add_rec = json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8"))
        r = self._add(led, add_rec)
        opp_id = r["opp_id"]
        cust_id = r["customer_id"]

        # 记一次跟进（逾期复跟推算）
        log_res = self._add(led, {"intent": "log_followup", "match": {"opp_id": opp_id},
                                  "follow_up": {"content": "客户说本周内给答复", "followed_at": "2026-08-06"}})
        self.assertEqual(log_res["opp_id"], opp_id)
        self.assertEqual(len(led["follow_log"]), 2)
        self.assertEqual(crm.find_opp_by_id(led, opp_id)["latest_follow"], "客户说本周内给答复")

        # 推进到谈判中，再签约（V1.2：签约不需金额，关闭普通回访）
        self._add(led, {"intent": "update_opportunity", "match": {"opp_id": opp_id},
                        "opportunity": {"stage": "谈判中"}})
        close_res = self._add(led, {"intent": "close_deal", "match": {"opp_id": opp_id},
                                    "opportunity": {"stage": "已签约"}})
        self.assertEqual(close_res["stage"], "已签约")
        opp = crm.find_opp_by_id(led, opp_id)
        self.assertIsNone(opp["next_follow_date"])

        # 回款（必填 payment_date + payment_amount）
        pay_res = self._add(led, {"intent": "update_payment", "match": {"opp_id": opp_id},
                                  "opportunity": {"payment_date": "2026-08-08", "payment_amount": 45000}})
        self.assertEqual(pay_res["stage"], "已回款")
        opp = crm.find_opp_by_id(led, opp_id)
        self.assertEqual(opp["payment_amount"], 45000)
        self.assertEqual(opp["payment_date"], "2026-08-08")

        # 汇总：签约=1（累计里程碑），回款=1
        s = crm._aggregate(led, RULES, "2026-08-08", "monthly")
        self.assertEqual(s["signed_count"], 1)
        self.assertEqual(s["collected_count"], 1)
        self.assertEqual(s["collected_amount"], 45000)

        # 客户档案状态变活跃 + 档案汇总
        cust = crm.find_customer_by_id(led, cust_id)
        self.assertEqual(cust["customer_status"], "活跃")
        agg = crm._customer_aggregates(led, cust_id, 2026)
        self.assertEqual(agg["history_deal_count"], 1)
        self.assertEqual(agg["total_deal_amount"], 45000)
        self.assertEqual(agg["first_coop_date"], "2026-08-08")

    def test_illegal_transition_rejected(self):
        led = fresh_ledger()
        self._add(led, json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8")))
        # 方案报价 不能跳回 初次接触
        res = crm.cmd_update(args_for(["--mode", "update", "--opp-id", "MO-2026-0001", "--stage", "初次接触"]),
                             led, RULES, TODAY)
        self.assertEqual(res["error"], "illegal_transition")

    def test_close_lost_requires_reason(self):
        led = fresh_ledger()
        self._add(led, json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8")))
        res = crm.cmd_close(args_for(["--mode", "close", "--opp-id", "MO-2026-0001", "--stage", "已流失"]),
                            led, RULES, TODAY)
        self.assertEqual(res["error"], "missing_field")


class TestDue(unittest.TestCase):
    def test_overdue_today_upcoming_nodate(self):
        led = fresh_ledger()
        led["opportunities"] = [
            {"opp_id": "MO-1", "customer_name": "逾期", "stage": "谈判中", "intent_level": "A",
             "owner": "张", "store": "S", "next_follow_date": "2026-07-20"},
            {"opp_id": "MO-2", "customer_name": "今日", "stage": "方案报价", "intent_level": "B",
             "owner": "张", "store": "S", "next_follow_date": TODAY},
            {"opp_id": "MO-3", "customer_name": "远期", "stage": "谈判中", "intent_level": "C",
             "owner": "张", "store": "S", "next_follow_date": "2026-09-30"},
            {"opp_id": "MO-4", "customer_name": "断档", "stage": "需求确认", "intent_level": "D",
             "owner": "张", "store": "S", "next_follow_date": None},
            {"opp_id": "MO-5", "customer_name": "已签约", "stage": "已签约", "intent_level": "A",
             "owner": "张", "store": "S", "next_follow_date": "2026-07-01"},  # 终态不提醒
        ]
        res = crm.cmd_due(args_for(["--mode", "due"]), led, RULES, TODAY)
        self.assertEqual(res["counts"]["overdue"], 1)
        self.assertEqual(res["counts"]["today"], 1)
        self.assertEqual(res["counts"]["upcoming"], 0)   # 9-30 超出7天窗口
        self.assertEqual(res["counts"]["no_date"], 1)
        self.assertEqual(res["overdue"][0]["opp_id"], "MO-1")
        self.assertIn("script", res["overdue"][0])


class TestExportXlsx(unittest.TestCase):
    def test_export_matches_template_layout(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError:
            self.skipTest("openpyxl not installed")
        led = fresh_ledger()
        crm.apply_record(led, RULES, TODAY, json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8")), confirmed=True)
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "export.xlsx"
            res = crm.cmd_export_xlsx(args_for(["--mode", "export-xlsx", "--out", str(out)]), led, RULES, TODAY)
            self.assertTrue(out.exists(), res)
            wb = openpyxl.load_workbook(out, data_only=True)
            self.assertEqual(wb.sheetnames, ["门店客户商机表", "总部客户档案表"])
            ws = wb["门店客户商机表"]
            # 表头在第4行，列数=20，首列=商机编号
            headers = [ws.cell(row=4, column=c).value for c in range(1, 23)]
            self.assertEqual(headers[0], "商机编号")
            self.assertEqual(headers[12], "预计金额")
            # 数据从第5行
            self.assertEqual(ws.cell(row=5, column=1).value, "MO-2026-0001")
            self.assertEqual(ws.cell(row=5, column=14).value, "方案报价")
            self.assertEqual(ws.cell(row=5, column=13).value, 45000)
            ws2 = wb["总部客户档案表"]
            self.assertEqual(ws2.cell(row=5, column=1).value, "KH-2026-0001")


class TestQuoteDelegation(unittest.TestCase):
    def test_dry_run_builds_v6_command_and_derives_budget(self):
        a = args_for(["--mode", "quote", "--customer", "八方科技", "--amount", "15000",
                      "--quantity", "300", "--order-quantity", "300", "--notes", "员工福利",
                      "--delivery-area", "长沙市内", "--list-plans", "3", "--dry-run"])
        res = crm.cmd_quote(a, fresh_ledger(), RULES, TODAY)
        self.assertNotIn("error", res)
        cmd = res["cmd"]
        self.assertEqual(res["budget_per_box"], 50.0)  # 15000/300
        self.assertIn("generate_excel_quote.mjs", cmd[1])
        self.assertIn("--budget-per-box", cmd)
        self.assertIn("50.0", cmd)
        self.assertIn("--list-plans", cmd)
        self.assertNotIn("--profit-margin", cmd)

    def test_quote_requires_budget_or_derivation(self):
        a = args_for(["--mode", "quote", "--customer", "X", "--order-quantity", "100"])
        res = crm.cmd_quote(a, fresh_ledger(), RULES, TODAY)
        self.assertEqual(res["error"], "missing_field")


class TestReviewReport(unittest.TestCase):
    def test_gen_report_renders_html(self):
        led = fresh_ledger()
        crm.apply_record(led, RULES, TODAY, json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8")), confirmed=True)
        summary = crm._aggregate(led, RULES, TODAY, "weekly")
        due = crm.cmd_due(args_for(["--mode", "due"]), led, RULES, TODAY)
        gen = _load("gen_under_test", "gen_review_report.py")
        with tempfile.TemporaryDirectory() as d:
            out = Path(d) / "review.html"
            payload_file = Path(d) / "payload.json"
            payload_file.write_text(json.dumps({"summary": summary, "due": due}, ensure_ascii=False), encoding="utf-8")
            rc = gen.main(["--summary", str(payload_file), "--out", str(out), "--title", "体验中心店"])
            self.assertEqual(rc, 0)
            html = out.read_text(encoding="utf-8")
            self.assertIn("__REPORT_DATA__", (SKILL_DIR / "assets" / "review-template.html").read_text(encoding="utf-8"))
            self.assertNotIn("__REPORT_DATA__", html)  # 占位符已替换
            self.assertIn("团购复盘周报", html)
            self.assertIn("深圳八方科技有限公司", html) or self.assertIn("45000", html)


class TestReliability(unittest.TestCase):
    """TC-096/097/098：事务一致性 / 乐观锁 / 幂等"""

    def _add(self, led):
        return crm.apply_record(led, RULES, TODAY,
                                json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8")),
                                confirmed=True)

    def test_tc096_tasks_lifecycle_transactional(self):
        led = fresh_ledger()
        r = self._add(led)
        opp_id = r["opp_id"]
        # 新建即有普通回访任务（与商机同事务）
        self.assertEqual([(t["type"], t["status"]) for t in led["tasks"]], [("普通回访", "待办")])
        # 推进到签约：回访关闭 + 回款待办
        for st in ["需求确认", "方案报价", "谈判中"]:
            crm.cmd_update(args_for(["--mode", "update", "--opp-id", opp_id, "--stage", st]), led, RULES, TODAY)
        crm.cmd_close(args_for(["--mode", "close", "--opp-id", opp_id, "--stage", "已签约"]), led, RULES, TODAY)
        ts = {t["type"]: t["status"] for t in led["tasks"]}
        self.assertEqual(ts["普通回访"], "已关闭")
        self.assertEqual(ts["回款"], "待办")
        # 回款：回款完成 + 交货待办
        crm.cmd_close(args_for(["--mode", "close", "--opp-id", opp_id, "--stage", "已回款",
                                "--payment-date", "2026-08-08", "--payment-amount", "45000", "--confirmed"]), led, RULES, TODAY)
        ts = {t["type"]: t["status"] for t in led["tasks"]}
        self.assertEqual(ts["回款"], "已完成")
        self.assertEqual(ts["交货"], "待办")
        # 交货日期写入：交货完成
        crm.cmd_update(args_for(["--mode", "update", "--opp-id", opp_id, "--delivery-date", "2026-08-20", "--confirmed"]), led, RULES, TODAY)
        self.assertEqual({t["type"]: t["status"] for t in led["tasks"]}["交货"], "已完成")

    def test_tc097_optimistic_lock(self):
        led = fresh_ledger()
        r = self._add(led)
        opp_id = r["opp_id"]
        res = crm.cmd_update(args_for(["--mode", "update", "--opp-id", opp_id,
                                       "--stage", "谈判中", "--expect-version", "999"]), led, RULES, TODAY)
        self.assertEqual(res["error"], "conflict")
        self.assertEqual(crm.find_opp_by_id(led, opp_id)["stage"], "方案报价")  # 未变
        ok = crm.cmd_update(args_for(["--mode", "update", "--opp-id", opp_id,
                                      "--stage", "谈判中", "--expect-version", str(crm.find_opp_by_id(led, opp_id)["version"])]),
                            led, RULES, TODAY)
        self.assertTrue(ok["_mutated"])
        self.assertEqual(crm.find_opp_by_id(led, opp_id)["stage"], "谈判中")

    def test_tc098_idempotent_and_dedup(self):
        led = fresh_ledger()
        rec = json.loads((TESTS / "fixture_valid.json").read_text(encoding="utf-8"))
        r1 = crm.apply_record(led, RULES, TODAY, rec, confirmed=True)
        dup = crm.cmd_log(args_for(["--mode", "log", "--opp-id", r1["opp_id"], "--content", "同一条跟进",
                                    "--followed-at", TODAY, "--confirmed"]), led, RULES, TODAY)
        dup2 = crm.cmd_log(args_for(["--mode", "log", "--opp-id", r1["opp_id"], "--content", "同一条跟进",
                                     "--followed-at", TODAY, "--confirmed"]), led, RULES, TODAY)
        self.assertEqual(dup2.get("status"), "duplicate")  # 同日同内容不重复写入
        self.assertEqual(len([f for f in led["follow_log"] if f["content"] == "同一条跟进"]), 1)


if __name__ == "__main__":
    unittest.main(verbosity=2)
