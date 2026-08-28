#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""门店团购客户与商机台账 · 确定性引擎（双表模型）。

数据模型对齐《月饼哥哥_团购客户管理双表模板.xlsx》：
  - 门店客户商机表：一条商机一行（含联系人/最近跟进/成交/流失内联字段）
  - 总部客户档案表：一家客户一行（含汇总与总部维护字段）
  - 引擎内部另有 follow_log（跟进历史）与 payment_status（回款分层，非双表列）

范式照 mooncake-ecommerce-analytics：扁平 argparse --mode + mode_map + --json；
每个模式返回 dict；致命错 stderr+exit(1)，数据缺失/需确认走 in-band {"error"/"status"}；
台账原子写（.tmp → fsync → os.replace）。引擎绝不调用模型。
"""
import argparse
import json
import os
import re
import subprocess
import sys
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace

# ACT-008: 中国大陆手机号正则
PHONE_REGEX = re.compile(r'^1[3-9]\d{9}$')

SKILL_DIR = Path(__file__).resolve().parent.parent
RULES_PATH = SKILL_DIR / "references" / "crm-rules.json"
DEFAULT_LEDGER = SKILL_DIR / "data" / "ledger.json"

# ----------------------------- 本地主库 -----------------------------
# 本地 JSON 是唯一主库；飞书 Base 是显式同步的交付镜像，不是可替换的存储后端。
# 在线同步由 sync_feishu_bitable.py 负责，避免网络失败被静默伪装成本地写入成功。
def get_storage_backend(path=None):
    """返回本地主库实例。path 指定时使用该隔离路径。"""
    return LocalJsonBackend(path or DEFAULT_LEDGER)


class LocalJsonBackend:
    """本地JSON文件后端（默认，向后兼容）。"""
    def __init__(self, path):
        self.path = Path(path)
    
    def load(self):
        p = self.path
        if not p.exists():
            return empty_ledger()
        try:
            led = json.loads(p.read_text(encoding="utf-8"))
            led.setdefault("operation_log", [])
            led.setdefault("tasks", [])
            led.setdefault("idempotency", {})
            led["schema_version"] = 2
            return led
        except json.JSONDecodeError as e:
            raise SystemExit(f"ledger JSON 损坏: {e}")
    
    def save(self, ledger):
        p = self.path
        p.parent.mkdir(parents=True, exist_ok=True)
        tmp = p.with_suffix(".tmp")
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(ledger, f, ensure_ascii=False, indent=2, default=str)
            f.flush()
            os.fsync(f.fileno())
        os.replace(tmp, p)
    
    def ensure_done(self):
        done = self.path.parent / ".done"
        done.touch()


PIPELINE_STAGES = ["初次接触", "需求确认", "方案报价", "谈判中"]
EXTRA_STORES = ["市场部"]

FOLLOW_SCRIPTS = {
    "初次接触": "您好{contact}，我是{store}的{owner}。之前您咨询的团购还考虑吗？方便的话我先发份方案给您参考。",
    "需求确认": "{contact}您好，跟您确认下：{product}约{quantity}盒，用途是{need}。我按这个给您核个准确的价格。",
    "方案报价": "{contact}您好，{product}{quantity}盒的报价方案您看了吗？这边走量有优惠，本周定能赶上档期，我帮您留货。",
    "谈判中": "{contact}您好，您内部定得怎么样了？需要补样品或调整方案随时说，我配合您。",
}


# ----------------------------- 基础工具 -----------------------------

def load_rules(path=RULES_PATH):
    rules = json.loads(Path(path).read_text(encoding="utf-8"))
    stores = rules.setdefault("stores", [])
    for store in EXTRA_STORES:
        if store not in stores:
            stores.append(store)
    return rules


def empty_ledger():
    return {"schema_version": 2, "customers": [], "opportunities": [], "follow_log": [], "operation_log": [], "tasks": [], "idempotency": {}, "seq": {}}


def load_ledger(path=None):
    backend = get_storage_backend(path)
    return backend.load()


def save_ledger(path=None, ledger=None):
    # 兼容旧签名：save_ledger(path, ledger)
    if ledger is None and isinstance(path, dict):
        ledger = path
        path = None
    backend = get_storage_backend(path)
    backend.save(ledger)


def parse_date(value):
    if not value:
        return None
    s = str(value).strip()[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


# ----------------------------- ACT-006/007/008: 输入校验 -----------------------------

def validate_date_format(value, field_name):
    """ACT-006: 校验日期是否为合法 ISO 8601 格式 (YYYY-MM-DD)。
    非 ISO 格式（如"下周三"）会被拒绝，返回结构化错误。
    """
    if not value:
        return None
    s = str(value).strip()
    try:
        datetime.strptime(s[:10], "%Y-%m-%d")
        return None  # 合法
    except ValueError:
        return {"error": "validation_failed", "field": field_name,
                "detail": f"日期格式错误: '{value}'，应为 ISO 8601 格式 (YYYY-MM-DD)，如 2026-08-10"}


def validate_non_negative(value, field_name):
    """ACT-007: 校验金额字段不能为负数。"""
    if value is None:
        return None
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None  # 非数字交给后续逻辑处理
    if v < 0:
        return {"error": "validation_failed", "field": field_name,
                "detail": f"{field_name}不能为负数，收到: {value}"}
    return None


def validate_positive(value, field_name):
    """ACT-007 + V1.2: 数量必须为正整数（拒绝小数/负数/文本）。"""
    if value is None:
        return None
    if isinstance(value, bool):
        return {"error": "validation_failed", "field": field_name,
                "detail": f"{field_name}必须为正整数，收到: {value}"}
    try:
        if isinstance(value, str):
            s = value.strip()
            if not s.lstrip("-").isdigit():
                return {"error": "validation_failed", "field": field_name,
                        "detail": f"{field_name}必须为正整数，收到: {value}"}
            v = int(s)
        elif isinstance(value, float):
            if not value.is_integer():
                return {"error": "validation_failed", "field": field_name,
                        "detail": f"{field_name}必须为正整数，收到小数: {value}"}
            v = int(value)
        else:
            v = int(value)
    except (TypeError, ValueError):
        return {"error": "validation_failed", "field": field_name,
                "detail": f"{field_name}必须为正整数，收到: {value}"}
    if v <= 0:
        return {"error": "validation_failed", "field": field_name,
                "detail": f"{field_name}必须大于 0，收到: {value}"}
    return None


def validate_phone(value):
    """ACT-008 + V1.2：联系电话/微信为文本字段。
    11位纯数字按大陆手机号格式校验；微信号/带前导零号码等非11位纯数字原样保留。"""
    if not value:
        return None
    s = str(value).strip()
    if s.isdigit() and len(s) == 11 and not PHONE_REGEX.match(s):
        return {"error": "validation_failed", "field": "contact_phone",
                "detail": f"手机号格式错误: '{value}'，应为中国大陆手机号 (1开头，第二位3-9，共11位)；若为微信号请直接说明"}
    return None


def validate_inputs(args, check_fields=None):
    """统一校验入口。check_fields 指定要检查的字段类别列表。
    返回 error dict 或 None（全部通过）。
    """
    if check_fields is None:
        check_fields = ["date", "amount", "quantity", "phone"]
    
    # ACT-006: 日期字段校验
    if "date" in check_fields:
        for attr, label in [("next_date", "下次跟进日期"), ("followed_at", "跟进日期"),
                             ("deal_date", "成交日期"), ("payment_date", "回款日期"),
                             ("delivery_date", "交货日期")]:
            val = getattr(args, attr, None)
            if val:
                err = validate_date_format(val, label)
                if err:
                    return err
    
    # ACT-007: 金额非负校验
    if "amount" in check_fields:
        val = getattr(args, "amount", None)
        if val is not None:
            err = validate_non_negative(val, "金额")
            if err:
                return err
        deal_val = getattr(args, "deal_amount", None)
        if deal_val is not None:
            err = validate_non_negative(deal_val, "成交金额")
            if err:
                return err
        pay_val = getattr(args, "payment_amount", None)
        if pay_val is not None:
            err = validate_non_negative(pay_val, "回款金额")
            if err:
                return err
    
    # ACT-007: 数量正数校验
    if "quantity" in check_fields:
        val = getattr(args, "quantity", None)
        if val is not None:
            err = validate_positive(val, "数量")
            if err:
                return err
    
    # ACT-008: 手机号正则校验
    if "phone" in check_fields:
        val = getattr(args, "contact_phone", None)
        if val:
            err = validate_phone(val)
            if err:
                return err
    
    return None


def to_iso(value):
    d = parse_date(value)
    return d.isoformat() if d else None


def next_id(ledger, prefix, year):
    key = f"{prefix}-{year}"
    n = int(ledger["seq"].get(key, 0)) + 1
    ledger["seq"][key] = n
    return f"{prefix}-{year}-{n:04d}"


def norm_name(name):
    return (name or "").strip().casefold()


def sync_tasks(ledger, rules, opp, today):
    """TC-096：提醒任务实体化，与商机变更同一原子写（tmp+rename），保证事务一致性。
    任务类型：普通回访 / 回款 / 交货。同opp同type未关闭任务只留一条。"""
    tasks = ledger.setdefault("tasks", [])
    stops = stop_states(rules)

    def live(opp_id, ttype):
        return [t for t in tasks if t["opp_id"] == opp_id and t["type"] == ttype
                and t["status"] != "已关闭"]

    # 普通回访：有next_follow_date且非停止态 → 建/更新日期；否则关闭
    if opp.get("next_follow_date") and opp.get("stage") not in stops:
        existing = live(opp["opp_id"], "普通回访")
        if existing:
            existing[0]["due_date"] = opp["next_follow_date"]
        else:
            tasks.append({"task_id": next_id(ledger, "TK", str(parse_date(today).year)),
                          "opp_id": opp["opp_id"], "type": "普通回访",
                          "due_date": opp["next_follow_date"], "status": "待办",
                          "created_at": today})
    else:
        for t in live(opp["opp_id"], "普通回访"):
            t["status"] = "已关闭"
    # 回款任务：已签约未回款 → 建/留待办
    if opp.get("stage") == "已签约" and not opp.get("payment_date"):
        if not live(opp["opp_id"], "回款"):
            tasks.append({"task_id": next_id(ledger, "TK", str(parse_date(today).year)),
                          "opp_id": opp["opp_id"], "type": "回款",
                          "due_date": None, "status": "待办", "created_at": today})
    elif opp.get("stage") == "已回款":
        for t in live(opp["opp_id"], "回款"):
            t["status"] = "已完成"
    # 交货任务：已回款未交货 → 建待办；有交货日期 → 已完成
    if opp.get("stage") == "已回款":
        if opp.get("delivery_date"):
            for t in live(opp["opp_id"], "交货"):
                t["status"] = "已完成"
                t["due_date"] = t.get("due_date") or opp["delivery_date"]
        else:
            existing = live(opp["opp_id"], "交货")
            if not existing:
                tasks.append({"task_id": next_id(ledger, "TK", str(parse_date(today).year)),
                              "opp_id": opp["opp_id"], "type": "交货",
                              "due_date": None, "status": "待办", "created_at": today})
    # 已流失：全部关闭
    if opp.get("stage") == "已流失":
        for t in live(opp["opp_id"], "普通回访") + live(opp["opp_id"], "回款") + live(opp["opp_id"], "交货"):
            t["status"] = "已关闭"


def normalize_enum(rules, value, list_key, alias_key):
    """V1.2：门店/线索来源写入前归一化（别名→标准枚举）。"""
    if not value:
        return value
    v = str(value).strip()
    return rules.get(alias_key, {}).get(v, v)


def stage_history_contains(opp, stage):
    """累计里程碑口径：stage_history 到过该阶段，或当前即该阶段。"""
    if opp.get("stage") == stage:
        return True
    return any(h.get("stage") == stage for h in (opp.get("stage_history") or []))


def compute_next_follow_date(rules, stage, intent_level, from_date, overdue=False):
    rd = rules["reminder_days"]
    if overdue:
        days = rd.get("overdue_retry", 1)
        return (from_date + timedelta(days=days)).isoformat()
    after = rd.get("after_stage", {})
    if stage in after and after[stage] is not None:
        days = after[stage]
    else:
        days = rd.get("by_intent_level", {}).get(intent_level)
    if days is None:
        return None
    return (from_date + timedelta(days=days)).isoformat()


def transition_allowed(rules, current, target):
    if current == target:
        return True
    machine = rules["status_machine"]["transitions"]
    return target in machine.get(current, [])


def stop_states(rules):
    sm = rules["status_machine"]
    return set(sm.get("reminder_stop_states", sm.get("terminal_states", [])))


def follow_script(stage, opp):
    tpl = FOLLOW_SCRIPTS.get(stage, "{contact}您好，我是{store}的{owner}，跟进下您之前提到的团购需求，方便聊两句吗？")
    return tpl.format(
        contact=opp.get("contact") or "您",
        store=opp.get("store") or "门店",
        owner=opp.get("owner") or "店长",
        product=opp.get("product") or "礼盒",
        quantity=opp.get("quantity") or "",
        need=opp.get("customer_need") or "团购",
    )


# ----------------------------- 客户/商机定位 -----------------------------

def find_customer_by_id(ledger, cust_id):
    for c in ledger["customers"]:
        if c.get("cust_id") == cust_id:
            return c
    return None


def find_customers_by_name(ledger, name):
    target = norm_name(name)
    return [c for c in ledger["customers"] if norm_name(c.get("name")) == target]


def find_opp_by_id(ledger, opp_id):
    for o in ledger["opportunities"]:
        if o.get("opp_id") == opp_id:
            return o
    return None


def resolve_opp(ledger, match):
    """按 match 定位唯一商机。返回 (opp, error_dict_or_None)。"""
    opp_id = match.get("opp_id")
    if opp_id:
        opp = find_opp_by_id(ledger, opp_id)
        return (opp, None) if opp else (None, {"error": "not_found", "detail": f"找不到商机 {opp_id}"})
    cust_id = match.get("cust_id")
    if cust_id:
        opps = [o for o in ledger["opportunities"] if o.get("cust_id") == cust_id]
        if not opps:
            return None, {"error": "not_found", "detail": f"客户 {cust_id} 名下没有商机"}
        if len(opps) > 1:
            return None, {"error": "ambiguous", "detail": f"客户 {cust_id} 有 {len(opps)} 条商机，请用 opp_id 指定",
                          "candidates": [o["opp_id"] for o in opps]}
        return opps[0], None
    name = match.get("customer_name")
    phone = match.get("contact_phone")
    if name or phone:
        opps = ledger["opportunities"]
        if name:
            opps = [o for o in opps if norm_name(o.get("customer_name")) == norm_name(name)]
        if phone:
            opps = [o for o in opps if (o.get("contact_phone") or "") == phone]
        if not opps:
            return None, {"error": "not_found", "detail": "按名称/电话找不到商机"}
        if len(opps) > 1:
            return None, {"error": "ambiguous", "detail": f"匹配到 {len(opps)} 条商机，请用 opp_id 指定",
                          "candidates": [o["opp_id"] for o in opps]}
        return opps[0], None
    return None, {"error": "missing_match", "detail": "未提供定位目标（opp_id/cust_id/customer_name/contact_phone）"}


# ----------------------------- 模式实现 -----------------------------

def confirm_gate(a):
    """关键控制（模板·语音字段标注版）：金额/数量/电话/日期须确认。
    返回本次写入命中的待确认字段标签清单（仅统计用户显式给出的值；系统推算的下次跟进日期不算）。"""
    hits = []
    if getattr(a, "amount", None) is not None:
        hits.append("金额")
    if getattr(a, "payment_amount", None) is not None:
        hits.append("回款金额")
    if getattr(a, "quantity", None) is not None:
        hits.append("数量")
    if getattr(a, "contact_phone", None):
        hits.append("电话")
    for attr, label in (("next_date", "下次跟进日期"), ("deal_date", "成交日期"), ("followed_at", "跟进日期"),
                        ("payment_date", "回款日期"), ("delivery_date", "交货日期")):
        if getattr(a, attr, None):
            hits.append(label)
    return hits


def confirm_refusal(hits):
    return {"status": "needs_confirmation", "reason": "confirm_required", "fields": hits,
            "detail": "以下字段需店长确认后写入：" + "、".join(hits) + "。确认无误后加 --confirmed 重新执行。",
            "_human": "待确认：" + "、".join(hits) + "。请向店长逐一回读确认，确认后加 --confirmed 写入。"}


def cmd_add(a, ledger, rules, today):
    if not a.name:
        return {"error": "missing_field", "detail": "新增商机必须有客户名称（--name）"}
    # V1.2 必填校验（对应模板带*字段）
    if not a.contact:
        return {"error": "missing_field", "detail": "新增商机必须有联系人（--contact）；缺失应先追问"}
    if not a.contact_phone:
        return {"error": "missing_field", "detail": "新增商机必须有联系电话/微信（--contact-phone）；缺失应先追问"}
    if not a.store:
        return {"error": "missing_field", "detail": "新增商机必须有所属门店（--store）"}
    if not a.need:
        return {"error": "missing_field", "detail": "新增商机必须有客户需求（--need）；缺失应先追问"}
    # V1.2.4：公司地址/线索来源/需求类别升级为必填（杰哥 08-11 拍板）
    if not a.customer_address:
        return {"error": "missing_field", "detail": "新增商机必须有公司地址（--customer-address；联系人所在地址，非配送地址）；缺失应先追问"}
    if not a.source:
        return {"error": "missing_field", "detail": "新增商机必须有线索来源（--source：新获客/老客转介绍/老客续费）；缺失应先追问，不要猜"}
    if not a.demand_category:
        return {"error": "missing_field",
                "detail": "新增商机必须有需求类别（--demand-category：散装/68礼盒/98礼盒/138礼盒/定制礼盒/待定）；判定规则见 SKILL.md，禁止用客户原话自由填写"}
    if a.demand_category and a.demand_category not in rules["demand_categories"]:
        return {"error": "bad_enum", "detail": f"需求类别 {a.demand_category} 非法，应为 {rules['demand_categories']}"}
    if a.amount is not None:
        a.amount = round(a.amount, 2)  # V1.2：金额保留两位小数
    if not a.intent_level:
        return {"error": "missing_field",
                "detail": "新增商机必须明确意向等级 A/B/C/D（--intent-level）；话语里没说应先追问，不要猜"}
    if a.intent_level not in rules["intent_levels"]:
        return {"error": "bad_enum", "detail": f"意向等级 {a.intent_level} 非法，应为 {rules['intent_levels']}"}
    stage = a.stage or rules["status_machine"]["default_entry_stage"]
    if stage not in rules["status_machine"]["states"]:
        return {"error": "bad_enum", "detail": f"当前阶段 {stage} 非法"}

    # ACT-006/007/008: 输入格式校验
    validation_err = validate_inputs(a, check_fields=["date", "amount", "quantity", "phone"])
    if validation_err:
        return validation_err

    if not a.confirmed:  # 关键控制硬闸口：金额/数量/电话/日期须确认
        hits = confirm_gate(a)
        if hits:
            return confirm_refusal(hits)

    year = str(parse_date(today).year)
    # 去重：按名称找客户
    existing = find_customers_by_name(ledger, a.name)
    phone = a.contact_phone
    customer = None
    if existing:
        cand = existing[0]
        cand_phone = cand.get("contact_phone")
        if phone and cand_phone and phone != cand_phone:
            return {"status": "needs_confirmation", "reason": "phone_mismatch",
                    "detail": f"已有同名客户 {cand['cust_id']}（电话 {cand_phone}），本次电话 {phone} 不同。"
                              f"确认是同一客户请加 --cust-id {cand['cust_id']}；确认是新客户请加 --new-customer。",
                    "existing": {"cust_id": cand["cust_id"], "name": cand["name"], "contact_phone": cand_phone}}
        customer = cand
    if a.cust_id:  # 显式挂到指定客户
        customer = find_customer_by_id(ledger, a.cust_id)
        if not customer:
            return {"error": "not_found", "detail": f"指定的客户 {a.cust_id} 不存在"}
    if customer is None and a.new_customer and existing:
        customer = None  # 强制新建

    # V1.2：门店/线索来源归一化（别名→标准枚举）
    a.store = normalize_enum(rules, a.store, "stores", "store_aliases")
    a.source = normalize_enum(rules, a.source, "lead_sources", "lead_source_aliases")
    if a.store and a.store not in rules["stores"]:
        return {"error": "bad_enum", "detail": f"所属门店 {a.store} 非法，应为 {rules['stores']}"}
    if a.source and a.source not in rules["lead_sources"]:
        return {"error": "bad_enum", "detail": f"线索来源 {a.source} 非法，应为 {rules['lead_sources']}"}

    if customer is None:
        cust_id = next_id(ledger, "KH", year)
        customer = {
            "cust_id": cust_id, "name": a.name.strip(), "short_name": a.short_name,
            "contact": a.contact, "contact_phone": phone,
            "customer_type": a.customer_type, "industry": a.industry,
            "customer_level": a.customer_level, "region": a.region, "address": a.address,
            "title": a.title, "other_contacts": a.other_contacts,
            "first_lead_source": a.source, "first_store": a.store, "current_store": a.store,
            "customer_owner": a.owner, "preference": a.preference, "tags": a.tags,
            "customer_status": "待激活", "notes": a.notes,
            "created_at": today, "archive_updated_at": today,
        }
        # V1.2：公司地址（联系人地址，非配送地址）优先写客户档案 address
        if a.customer_address and not customer.get("address"):
            customer["address"] = a.customer_address
        ledger["customers"].append(customer)
        created_customer = True
    else:
        created_customer = False
        # 用新信息补全空字段
        for field, val in [("contact", a.contact), ("contact_phone", phone), ("short_name", a.short_name),
                           ("customer_type", a.customer_type), ("industry", a.industry),
                           ("customer_level", a.customer_level), ("region", a.region), ("address", a.address),
                           ("title", a.title), ("other_contacts", a.other_contacts),
                           ("preference", a.preference), ("tags", a.tags), ("notes", a.notes)]:
            if val and not customer.get(field):
                customer[field] = val
        if a.store:
            customer["current_store"] = a.store
        if a.owner:
            customer["customer_owner"] = a.owner
        customer["archive_updated_at"] = today

    # 新建商机
    opp_id = next_id(ledger, "MO", year)
    # V1.2.4：提醒严格按 ABCD 天数推算，客户/店长口述时间不覆盖 ABCD 提醒；
    # 仅当等级规则无天数（D级）且口述了时间时，才按口述时间排，避免提醒丢失
    nd = compute_next_follow_date(rules, stage, a.intent_level, parse_date(today))
    if nd is None and a.next_date:
        nd = to_iso(a.next_date)
    opp = {
        "opp_id": opp_id, "cust_id": customer["cust_id"],
        "customer_name": customer["name"], "customer_address": a.customer_address,
        "contact": a.contact or customer.get("contact"),
        "contact_phone": phone or customer.get("contact_phone"),
        "store": a.store, "lead_source": a.source, "customer_need": a.need,
        "product": a.product, "demand_category": a.demand_category,
        "quantity": a.quantity, "intent_level": a.intent_level,
        "expected_amount": a.amount, "stage": stage, "latest_follow": a.follow_content,
        "next_follow_date": nd, "owner": a.owner,
        "payment_date": None, "payment_amount": None, "delivery_date": None,
        "lost_reason": None,
        "stage_history": [{"stage": stage, "at": today}],
        "version": 1,
        "created_at": today, "updated_at": today,
    }
    ledger["opportunities"].append(opp)

    fu_id = None
    if a.follow_content:
        fu_id = next_id(ledger, "FU", year)
        ledger["follow_log"].append({
            "fu_id": fu_id, "opp_id": opp_id, "cust_id": customer["cust_id"],
            "followed_at": to_iso(a.followed_at) or today, "content": a.follow_content,
            "customer_feedback": a.feedback, "next_action": a.next_action,
        })

    sync_tasks(ledger, rules, opp, today)  # TC-096：任务与商机同事务
    human = (f"已记录：客户「{customer['name']}」（{customer['cust_id']}，"
             f"{'新建' if created_customer else '已有'}）新增商机 {opp_id}，"
             f"意向 {a.intent_level} 级，阶段「{stage}」，下次跟进 {nd or '未排'}。")
    return {"_mutated": True, "_human": human, "customer_id": customer["cust_id"],
            "opp_id": opp_id, "follow_up_id": fu_id, "stage": stage,
            "intent_level": a.intent_level, "next_follow_date": nd,
            "version": opp["version"], "created_customer": created_customer}


def cmd_update(a, ledger, rules, today):
    opp, err = resolve_opp(ledger, _match(a))
    if err:
        return err
    # TC-097：乐观锁——期望版本不匹配则拒绝，不静默覆盖
    if getattr(a, "expect_version", None) is not None and opp.get("version", 1) != a.expect_version:
        return {"error": "conflict",
                "detail": f"商机 {opp['opp_id']} 已被其他人修改（当前版本 {opp.get('version', 1)}，"
                          f"你基于版本 {a.expect_version}），请重新查询后再改",
                "current_version": opp.get("version", 1)}

    # ACT-006/007/008: 输入格式校验
    validation_err = validate_inputs(a, check_fields=["date", "amount", "quantity", "phone"])
    if validation_err:
        return validation_err

    changes = {}
    if a.stage:
        a.stage = rules["status_machine"].get("stage_aliases", {}).get(a.stage, a.stage)
        if a.stage not in rules["status_machine"]["states"]:
            return {"error": "bad_enum", "detail": f"当前阶段 {a.stage} 非法"}
        if not transition_allowed(rules, opp.get("stage"), a.stage):
            return {"error": "illegal_transition",
                    "detail": f"阶段不能从「{opp.get('stage')}」直接跳到「{a.stage}」",
                    "allowed": rules["status_machine"]["transitions"].get(opp.get("stage"), [])}
        changes["stage"] = a.stage
    for src, dst in [("amount", "expected_amount"), ("quantity", "quantity"),
                     ("intent_level", "intent_level"), ("product", "product"),
                     ("need", "customer_need"), ("owner", "owner"),
                     ("contact", "contact"), ("contact_phone", "contact_phone"),
                     ("customer_address", "customer_address"),
                     ("demand_category", "demand_category"),
                     ("payment_date", "payment_date"),
                     ("payment_amount", "payment_amount"),
                     ("delivery_date", "delivery_date")]:
        val = getattr(a, src)
        if val is not None:
            if dst in ("expected_amount", "payment_amount") and isinstance(val, float):
                val = round(val, 2)
            changes[dst] = val
    if a.intent_level and a.intent_level not in rules["intent_levels"]:
        return {"error": "bad_enum", "detail": f"意向等级 {a.intent_level} 非法"}
    if a.demand_category and a.demand_category not in rules["demand_categories"]:
        return {"error": "bad_enum", "detail": f"需求类别 {a.demand_category} 非法，应为 {rules['demand_categories']}"}
    if a.store is not None:
        ns = normalize_enum(rules, a.store, "stores", "store_aliases")
        if ns not in rules["stores"]:
            return {"error": "bad_enum", "detail": f"所属门店 {ns} 非法，应为 {rules['stores']}"}
        changes["store"] = ns
    if a.next_date:
        # V1.2.4：严格按 ABCD 天数提醒，口述日期仅在等级规则无天数（D级）时兜底，不覆盖 ABCD 提醒
        new_stage = changes.get("stage", opp.get("stage"))
        new_level = changes.get("intent_level", opp.get("intent_level"))
        nd = compute_next_follow_date(rules, new_stage, new_level, parse_date(today))
        changes["next_follow_date"] = nd or to_iso(a.next_date)
    if not changes:
        return {"error": "no_change", "detail": "没有提供任何要更新的字段"}
    # V1.2：进入已回款必须齐备 回款日期+回款金额（已有或本次提供）
    if changes.get("stage") == "已回款":
        if not (changes.get("payment_date") or opp.get("payment_date")) or            (changes.get("payment_amount") is None and opp.get("payment_amount") is None):
            return {"error": "missing_field",
                    "detail": "进入「已回款」必须提供回款日期+回款金额（--payment-date/--payment-amount）"}
    if not a.confirmed:  # 关键控制硬闸口
        hits = confirm_gate(a)
        if hits:
            return confirm_refusal(hits)
    # V1.2：operation_log（门店/跟进人/金额/日期修改留痕）
    op_fields = set(rules.get("operation_log_fields", {}).get("fields", []))
    for field in op_fields & set(changes.keys()):
        old = opp.get(field)
        ledger.setdefault("operation_log", []).append(
            {"at": today, "opp_id": opp["opp_id"], "field": field, "old": old, "new": changes[field]})
    opp.update(changes)
    opp["updated_at"] = today
    if "stage" in changes:
        opp.setdefault("stage_history", []).append({"stage": changes["stage"], "at": today})
    opp["version"] = opp.get("version", 1) + 1  # TC-097
    sync_tasks(ledger, rules, opp, today)  # TC-096
    # 阶段或意向变化且未显式给下次日期时，重算提醒（停止态不排）
    if ("stage" in changes or "intent_level" in changes) and "next_follow_date" not in changes:
        if opp.get("stage") in stop_states(rules):
            opp["next_follow_date"] = None
        else:
            opp["next_follow_date"] = compute_next_follow_date(rules, opp.get("stage"), opp.get("intent_level"), parse_date(today))
    return {"_mutated": True, "_human": f"已更新商机 {opp['opp_id']}：{', '.join(changes.keys())}。",
            "opp_id": opp["opp_id"], "changes": changes, "version": opp["version"],
            "next_follow_date": opp.get("next_follow_date")}


def cmd_log(a, ledger, rules, today):
    opp, err = resolve_opp(ledger, _match(a))
    if err:
        return err
    if not a.content:
        return {"error": "missing_field", "detail": "记录跟进必须有内容（--content）"}
    # TC-097：乐观锁
    if getattr(a, "expect_version", None) is not None and opp.get("version", 1) != a.expect_version:
        return {"error": "conflict", "detail": f"商机已被修改（当前版本 {opp.get('version', 1)}）",
                "current_version": opp.get("version", 1)}
    # TC-098：同日同内容跟进去重（防重复提交）
    dup = [f for f in ledger["follow_log"]
           if f["opp_id"] == opp["opp_id"] and f.get("content") == a.content
           and f.get("followed_at") == (to_iso(a.followed_at) or today)]
    if dup:
        return {"status": "duplicate", "detail": "同一天的相同跟进记录已存在，未重复写入",
                "follow_up_id": dup[0]["fu_id"], "opp_id": opp["opp_id"]}

    # ACT-006: 日期格式校验
    validation_err = validate_inputs(a, check_fields=["date"])
    if validation_err:
        return validation_err

    if not a.confirmed:  # 关键控制硬闸口：显式日期须确认
        hits = confirm_gate(a)
        if hits:
            return confirm_refusal(hits)
    followed_at = to_iso(a.followed_at) or today
    year = str(parse_date(followed_at).year)
    fu_id = next_id(ledger, "FU", year)
    ledger["follow_log"].append({
        "fu_id": fu_id, "opp_id": opp["opp_id"], "cust_id": opp["cust_id"],
        "followed_at": followed_at, "content": a.content,
        "customer_feedback": a.feedback, "next_action": a.next_action,
    })
    opp["latest_follow"] = a.content
    if a.stage:
        a.stage = rules["status_machine"].get("stage_aliases", {}).get(a.stage, a.stage)
        if not transition_allowed(rules, opp.get("stage"), a.stage):
            return {"error": "illegal_transition",
                    "detail": f"阶段不能从「{opp.get('stage')}」跳到「{a.stage}」"}
        opp["stage"] = a.stage
        opp.setdefault("stage_history", []).append({"stage": a.stage, "at": followed_at})
    # V1.2.4 下次跟进：严格按 ABCD 天数提醒，客户口述时间不覆盖 ABCD 提醒（口述时间记入跟进内容）
    cur = parse_date(opp.get("next_follow_date"))
    overdue = bool(cur and cur < parse_date(followed_at))
    nd = compute_next_follow_date(rules, opp.get("stage"), opp.get("intent_level"), parse_date(followed_at), overdue=overdue)
    if nd is None and a.next_date:  # 等级规则无天数（D级）且口述了时间 → 按口述时间排，不丢提醒
        nd = to_iso(a.next_date)
    opp["next_follow_date"] = nd
    opp["updated_at"] = followed_at
    opp["version"] = opp.get("version", 1) + 1  # TC-097
    sync_tasks(ledger, rules, opp, today)  # TC-096
    return {"_mutated": True, "_human": f"已为 {opp['opp_id']} 记录跟进（{fu_id}），下次跟进 {opp['next_follow_date'] or '未排'}。",
            "opp_id": opp["opp_id"], "follow_up_id": fu_id, "stage": opp.get("stage"),
            "version": opp["version"], "next_follow_date": opp.get("next_follow_date")}


def cmd_close(a, ledger, rules, today):
    opp, err = resolve_opp(ledger, _match(a))
    if err:
        return err
    # TC-097：乐观锁
    if getattr(a, "expect_version", None) is not None and opp.get("version", 1) != a.expect_version:
        return {"error": "conflict", "detail": f"商机已被修改（当前版本 {opp.get('version', 1)}）",
                "current_version": opp.get("version", 1)}
    target = a.stage
    if target not in rules["status_machine"]["closing_states"]:
        return {"error": "bad_enum", "detail": f"close 的阶段只能是 {rules['status_machine']['closing_states']}，收到 {target}"}
    if not transition_allowed(rules, opp.get("stage"), target):
        return {"error": "illegal_transition",
                "detail": f"阶段不能从「{opp.get('stage')}」跳到「{target}」",
                "allowed": rules["status_machine"]["transitions"].get(opp.get("stage"), [])}

    # ACT-006/007: 日期和金额校验
    validation_err = validate_inputs(a, check_fields=["date", "amount"])
    if validation_err:
        return validation_err

    if not a.confirmed:  # 关键控制硬闸口：回款日期/回款金额须确认
        hits = confirm_gate(a)
        if hits:
            return confirm_refusal(hits)
    if target == "已签约":
        opp["stage"] = "已签约"
        opp["next_follow_date"] = None  # 关闭普通回访，回款任务独立
        cust = find_customer_by_id(ledger, opp["cust_id"])
        if cust:
            cust["customer_status"] = "活跃"
            cust["archive_updated_at"] = today
        human = f"商机 {opp['opp_id']} 已签约，普通回访关闭；待回款请记回款日期+回款金额。"
    elif target == "已回款":
        pay_date = to_iso(a.payment_date) or opp.get("payment_date")
        pay_amount = a.payment_amount if a.payment_amount is not None else opp.get("payment_amount")
        if not pay_date or pay_amount is None:
            return {"error": "missing_field",
                    "detail": "标记已回款必须提供回款日期+回款金额（--payment-date/--payment-amount）"}
        opp["stage"] = "已回款"
        opp["payment_date"] = pay_date
        opp["payment_amount"] = pay_amount
        # 冗余口径：供总部档案聚合（成交=回款口径）
        opp["deal_date"] = pay_date
        opp["deal_amount"] = pay_amount
        opp["payment_status"] = "已回款"
        opp["next_follow_date"] = None
        cust = find_customer_by_id(ledger, opp["cust_id"])
        if cust:
            cust["customer_status"] = "活跃"
            cust["archive_updated_at"] = today
        human = f"商机 {opp['opp_id']} 已回款 {pay_amount} 元（{pay_date}），交货日期待记。"
    else:  # 已流失
        reason = a.lost_reason or opp.get("lost_reason")
        if not reason:
            return {"error": "missing_field", "detail": "标记流失必须提供流失原因（--lost-reason）；缺失应先追问"}
        opp["stage"] = "已流失"
        opp["lost_reason"] = reason
        opp["next_follow_date"] = None
        cust = find_customer_by_id(ledger, opp["cust_id"])
        if cust:
            active = [o for o in ledger["opportunities"]
                      if o["cust_id"] == cust["cust_id"] and o["opp_id"] != opp["opp_id"]
                      and o.get("stage") in (PIPELINE_STAGES + ["已签约"])]
            if not active:
                cust["customer_status"] = "流失"
            cust["archive_updated_at"] = today
        human = f"商机 {opp['opp_id']} 已标记流失，原因：{reason}。"
    opp.setdefault("stage_history", []).append({"stage": target, "at": today})
    opp["updated_at"] = today
    opp["version"] = opp.get("version", 1) + 1  # TC-097
    sync_tasks(ledger, rules, opp, today)  # TC-096
    return {"_mutated": True, "_human": human, "opp_id": opp["opp_id"], "stage": target,
            "version": opp["version"],
            "payment_amount": opp.get("payment_amount"), "payment_date": opp.get("payment_date"),
            "lost_reason": opp.get("lost_reason")}


def cmd_show(a, ledger, rules, today):
    if a.opp_id:
        opp = find_opp_by_id(ledger, a.opp_id)
        if not opp:
            return {"error": "not_found", "detail": f"找不到商机 {a.opp_id}"}
        cust = find_customer_by_id(ledger, opp["cust_id"])
        fus = [f for f in ledger["follow_log"] if f["opp_id"] == opp["opp_id"]]
        return {"opportunity": opp, "customer": cust, "follow_log": fus}
    if a.cust_id:
        cust = find_customer_by_id(ledger, a.cust_id)
        if not cust:
            return {"error": "not_found", "detail": f"找不到客户 {a.cust_id}"}
        opps = [o for o in ledger["opportunities"] if o["cust_id"] == a.cust_id]
        fus = [f for f in ledger["follow_log"] if f["cust_id"] == a.cust_id]
        return {"customer": cust, "opportunities": opps, "follow_log": fus}
    return {"error": "missing_field", "detail": "show 需要 --opp-id 或 --cust-id"}


def cmd_due(a, ledger, rules, today):
    t = parse_date(today)
    window = rules["reminder_days"].get("upcoming_window_days", 7)
    stops = stop_states(rules)
    overdue, today_due, upcoming, no_date = [], [], [], []
    for o in ledger["opportunities"]:
        if o.get("stage") in stops:
            continue
        nd = parse_date(o.get("next_follow_date"))
        item = {"opp_id": o["opp_id"], "customer_name": o.get("customer_name"),
                "stage": o.get("stage"), "intent_level": o.get("intent_level"),
                "owner": o.get("owner"), "store": o.get("store"),
                "next_follow_date": o.get("next_follow_date"),
                "script": follow_script(o.get("stage"), o)}
        if nd is None:
            no_date.append(item)
        elif nd < t:
            item["days_overdue"] = (t - nd).days
            overdue.append(item)
        elif nd == t:
            today_due.append(item)
        elif (nd - t).days <= window:
            item["days_until"] = (nd - t).days
            upcoming.append(item)
    level_rank = {"A": 0, "B": 1, "C": 2, "D": 3}
    overdue.sort(key=lambda x: (level_rank.get(x["intent_level"], 9), -x.get("days_overdue", 0)))
    upcoming.sort(key=lambda x: (level_rank.get(x["intent_level"], 9), x.get("days_until", 99)))
    pending_tasks = [t for t in ledger.get("tasks", []) if t["status"] == "待办"]
    result = {"today": today, "window_days": window,
              "counts": {"overdue": len(overdue), "today": len(today_due),
                         "upcoming": len(upcoming), "no_date": len(no_date),
                         "pending_tasks": len(pending_tasks)},
              "overdue": overdue, "today_due": today_due, "upcoming": upcoming, "no_date": no_date,
              "pending_tasks": pending_tasks}
    lines = [f"截至 {today}：逾期 {len(overdue)}，今日 {len(today_due)}，未来{window}天 {len(upcoming)}，跟进断档 {len(no_date)}。"]
    for it in overdue[:5]:
        lines.append(f"  ⚠逾期{it['days_overdue']}天 {it['opp_id']} {it['customer_name']}（{it['intent_level']}级/{it['stage']}/{it['owner'] or '未指派'}）")
    result["_human"] = "\n".join(lines)
    return result


def cmd_query(a, ledger, rules, today):
    t = parse_date(today)
    stops = stop_states(rules)
    rows = ledger["opportunities"]
    if a.stage:
        rows = [o for o in rows if o.get("stage") == a.stage]
    if a.intent_level:
        rows = [o for o in rows if o.get("intent_level") == a.intent_level]
    if a.owner:
        rows = [o for o in rows if o.get("owner") == a.owner]
    if a.store:
        rows = [o for o in rows if o.get("store") == a.store]
    if a.cust_id:
        rows = [o for o in rows if o.get("cust_id") == a.cust_id]
    if a.overdue:
        rows = [o for o in rows if o.get("stage") not in stops
                and parse_date(o.get("next_follow_date")) is not None
                and parse_date(o.get("next_follow_date")) < t]
    out = [{"opp_id": o["opp_id"], "cust_id": o["cust_id"], "customer_name": o.get("customer_name"),
            "stage": o.get("stage"), "intent_level": o.get("intent_level"),
            "expected_amount": o.get("expected_amount"), "owner": o.get("owner"),
            "next_follow_date": o.get("next_follow_date")} for o in rows]
    return {"count": len(out), "opportunities": out, "_human": f"匹配 {len(out)} 条商机。"}


def _aggregate(ledger, rules, today, period):
    t = parse_date(today)
    span = {"daily": 0, "weekly": 6, "monthly": 29}.get(period, 6)
    start = t - timedelta(days=span)

    def in_window(iso):
        d = parse_date(iso)
        return bool(d and start <= d <= t)

    opps = ledger["opportunities"]
    custs = ledger["customers"]
    stops = stop_states(rules)
    pipeline = set(PIPELINE_STAGES)

    new_opps = [o for o in opps if in_window(o.get("created_at"))]
    new_custs = [c for c in custs if in_window(c.get("created_at"))]
    active = [o for o in opps if o.get("stage") not in stops]
    won = [o for o in opps if stage_history_contains(o, "已签约")]  # 累计里程碑口径
    lost = [o for o in opps if o.get("stage") == "已流失"]
    collected = [o for o in opps if o.get("stage") == "已回款"]

    intent_buckets = {lv: len([o for o in active if o.get("intent_level") == lv]) for lv in rules["intent_levels"]}
    stage_buckets = {s: len([o for o in opps if o.get("stage") == s]) for s in rules["status_machine"]["states"]}
    pipeline_amount = sum(o.get("expected_amount") or 0 for o in active if o.get("stage") in pipeline)
    won_amount = sum(o.get("payment_amount") or o.get("deal_amount") or 0 for o in collected)
    collected_amount = sum(o.get("payment_amount") or o.get("deal_amount") or 0 for o in collected)
    overdue = [o for o in active if parse_date(o.get("next_follow_date")) is not None
               and parse_date(o.get("next_follow_date")) < t]

    by_owner, by_store = {}, {}
    for o in active:
        for key, store in (("owner", by_owner), ("store", by_store)):
            k = o.get(key) or "未指派" if key == "owner" else (o.get(key) or "未知门店")
            e = store.setdefault(k, {"count": 0, "pipeline_amount": 0})
            e["count"] += 1
            if o.get("stage") in pipeline:
                e["pipeline_amount"] += o.get("expected_amount") or 0

    conv = None
    conv_note = None
    denom = len(won) + len(lost)
    if denom >= 5:
        conv = round(len(won) / denom, 3)
    else:
        conv_note = f"样本不足（成交{len(won)}+流失{len(lost)}<5），仅供参考"

    top_lost = [{"reason": r or "未记录", "count": c} for r, c in Counter(o.get("lost_reason") for o in lost).most_common(5)]

    level_rank = {"A": 0, "B": 1, "C": 2, "D": 3}
    focus = [o for o in active if parse_date(o.get("next_follow_date")) is not None
             and t <= parse_date(o.get("next_follow_date")) <= t + timedelta(days=7)]
    focus.sort(key=lambda o: (level_rank.get(o.get("intent_level"), 9), o.get("next_follow_date")))
    next_focus = [{"opp_id": o["opp_id"], "customer_name": o.get("customer_name"),
                   "intent_level": o.get("intent_level"), "stage": o.get("stage"),
                   "owner": o.get("owner"), "next_follow_date": o.get("next_follow_date")} for o in focus[:5]]

    return {
        "period": period, "window": {"start": start.isoformat(), "end": t.isoformat()},
        "new_opportunities": len(new_opps), "new_customers": len(new_custs),
        "intent_buckets": intent_buckets, "stage_buckets": stage_buckets,
        "signed_count": len(won), "won_count": len(collected),
        "collected_count": len(collected), "lost_count": len(lost),
        "pipeline_amount": pipeline_amount, "won_amount": won_amount, "collected_amount": collected_amount,
        "overdue_count": len(overdue), "by_owner": by_owner, "by_store": by_store,
        "conversion_rate": conv, "conversion_note": conv_note,
        "top_lost_reasons": top_lost, "next_week_focus": next_focus,
    }


def cmd_summary(a, ledger, rules, today):
    return _aggregate(ledger, rules, today, a.period)


def cmd_report(a, ledger, rules, today):
    """V1.2：按门店横向汇总报表（5门店+合计）。口径见 crm-rules.json report 段。"""
    opps = ledger["opportunities"]
    stores = list(rules["stores"])
    levels = rules["intent_levels"]

    def bucket(store_opps):
        total = len(store_opps)
        abcd = {lv: len([o for o in store_opps if o.get("intent_level") == lv]) for lv in levels}
        a_ratio = round(abcd["A"] / total * 100, 1) if total else 0.0
        est_amount = sum(o.get("expected_amount") or 0 for o in store_opps if o.get("stage") != "已流失")
        demand = Counter(o.get("demand_category") or "未分类" for o in store_opps)
        signed = len([o for o in store_opps if stage_history_contains(o, "已签约")])
        paid = len([o for o in store_opps if o.get("stage") == "已回款"])
        lost = len([o for o in store_opps if o.get("stage") == "已流失"])
        negotiating = len([o for o in store_opps if o.get("stage") == "谈判中"])
        conv = round(signed / total * 100, 1) if total else 0.0  # 0时显0%
        return {"团单总数": total, "A": abcd["A"], "B": abcd["B"], "C": abcd["C"], "D": abcd["D"],
                "A级占比": f"{a_ratio}%", "预计总金额": est_amount,
                "需求类别分布": dict(demand), "已签约数": signed, "已回款数": paid,
                "已流失数": lost, "谈判中数": negotiating, "签约转化率": f"{conv}%"}

    rows = []
    for st in stores:
        r = bucket([o for o in opps if o.get("store") == st])
        r["门店"] = st
        rows.append(r)
    total_row = bucket(opps)
    total_row["门店"] = "合计"
    rows.append(total_row)

    lines = [f"【按门店汇总报表 截至 {today}】"]
    lines.append("门店 | 总数 | A/B/C/D | A级占比 | 预计金额 | 已签约 | 已回款 | 已流失 | 谈判中 | 签约转化率")
    for r in rows:
        lines.append(f"{r['门店']} | {r['团单总数']} | {r['A']}/{r['B']}/{r['C']}/{r['D']} | "
                     f"{r['A级占比']} | {r['预计总金额']:.0f} | {r['已签约数']} | {r['已回款数']} | "
                     f"{r['已流失数']} | {r['谈判中数']} | {r['签约转化率']}")
    return {"today": today, "rows": rows, "_human": "\n".join(lines)}


def cmd_review(a, ledger, rules, today):
    s = _aggregate(ledger, rules, today, a.period)
    due = cmd_due(a, ledger, rules, today)
    cn = s["conversion_note"] or f"转化率 {s['conversion_rate']:.0%}"
    lines = []
    lines.append(f"【{s['period']}复盘 {s['window']['start']} ~ {s['window']['end']}】")
    lines.append(f"一句话结论：本期新增商机 {s['new_opportunities']} 条，在谈盘子 {s['pipeline_amount']:.0f} 元，"
                 f"已回款 {s['collected_amount']:.0f} 元，逾期 {s['overdue_count']} 条。")
    lines.append("")
    lines.append("漏斗快照：")
    lines.append(f"  阶段分布：{s['stage_buckets']}")
    lines.append(f"  意向分布（在谈）：{s['intent_buckets']}")
    lines.append(f"  在谈金额 {s['pipeline_amount']:.0f}；已签约(累计){s['signed_count']} 条，"
                 f"已回款 {s['collected_count']} 条/{s['collected_amount']:.0f} 元；流失 {s['lost_count']} 条。")
    lines.append(f"  {cn}")
    lines.append("")
    lines.append("风险清单（逾期/断档，点名到商机）：")
    risk = due["overdue"] + due["today_due"]
    if risk:
        for it in risk[:10]:
            tag = f"逾期{it.get('days_overdue', 0)}天" if "days_overdue" in it else "今日到期"
            lines.append(f"  {tag} {it['opp_id']} {it['customer_name']}（{it['intent_level']}级/{it['stage']}/{it['owner'] or '未指派'}）")
    else:
        lines.append("  无逾期。")
    if due["no_date"]:
        lines.append(f"  跟进断档（无下次跟进日期）：{', '.join(o['opp_id'] for o in due['no_date'][:10])}")
    lines.append("")
    lines.append("流失归因（证据等级见方法论，空缺需补录）：")
    if s["top_lost_reasons"]:
        for r in s["top_lost_reasons"]:
            lines.append(f"  {r['reason']}：{r['count']} 条")
    else:
        lines.append("  本期无流失。")
    lines.append("")
    lines.append("下周重点（高意向优先）：")
    if s["next_week_focus"]:
        for f in s["next_week_focus"]:
            lines.append(f"  {f['next_follow_date']} {f['opp_id']} {f['customer_name']}（{f['intent_level']}级/{f['stage']}）→ 按阶段话术跟进")
    else:
        lines.append("  未来7天无到期商机。")
    lines.append("")
    lines.append("数据缺口（请店长补齐，不要猜）：")
    gaps = []
    for o in ledger["opportunities"]:
        if o.get("stage") not in stop_states(rules) and not o.get("intent_level"):
            gaps.append(f"{o['opp_id']} 缺意向等级")
    if due["no_date"]:
        gaps.append(f"{len(due['no_date'])} 条缺下次跟进日期")
    lost_noreason = [o["opp_id"] for o in ledger["opportunities"] if o.get("stage") == "已流失" and not o.get("lost_reason")]
    if lost_noreason:
        gaps.append(f"流失无原因：{', '.join(lost_noreason[:5])}")
    lines.append("  " + ("；".join(gaps) if gaps else "无明显缺口。"))
    text = "\n".join(lines)
    return {"period": s["period"], "review_text": text, "summary": s,
            "due_counts": due["counts"], "_human": text}


def _customer_aggregates(ledger, cust_id, year):
    # V1.2：成交=回款口径；历史购买产品取 demand_category（回退 product）
    won = [o for o in ledger["opportunities"]
           if o.get("cust_id") == cust_id and stage_history_contains(o, "已签约")]
    dates = [parse_date(o.get("deal_date") or o.get("payment_date")) for o in won
             if parse_date(o.get("deal_date") or o.get("payment_date"))]
    products = []
    for o in won:
        p = o.get("demand_category") or o.get("product")
        if p and p not in products:
            products.append(p)
    amt = lambda o: o.get("payment_amount") or o.get("deal_amount") or 0
    return {
        "first_coop_date": min(dates).isoformat() if dates else None,
        "last_coop_date": max(dates).isoformat() if dates else None,
        "history_products": "、".join(products),
        "history_deal_count": len(won),
        "total_deal_amount": sum(amt(o) for o in won),
        "year_deal_amount": sum(amt(o) for o in won
                                if parse_date(o.get("deal_date") or o.get("payment_date"))
                                and parse_date(o.get("deal_date") or o.get("payment_date")).year == year),
    }


def build_sync_record_sets(ledger, rules, today):
    """生成飞书双表镜像的确定性 JSON 记录集，不访问网络。

    字段名、列序和主键均来自现有 export 契约，避免 Excel 成为在线同步的
    中间格式。返回值可直接交给 sync_feishu_bitable.py 做预览或写入。
    """
    year = parse_date(today).year
    exp = rules["export"]

    def make_table(table_name, primary_internal_key, source_rows):
        cfg = exp[table_name]
        columns = cfg["columns"]
        header_by_key = dict(columns)
        primary_key = header_by_key[primary_internal_key]
        records = []
        seen = set()
        errors = []
        for index, row_data in enumerate(source_rows, start=1):
            fields = {header: row_data.get(key) for key, header in columns}
            key_value = fields.get(primary_key)
            if key_value is None or str(key_value).strip() == "":
                errors.append({"type": "blank_primary_key", "row": index, "field": primary_key})
                continue
            key_text = str(key_value)
            if key_text in seen:
                errors.append({"type": "duplicate_primary_key", "row": index,
                               "field": primary_key, "value": key_text})
                continue
            seen.add(key_text)
            records.append({"key": key_text, "fields": fields})
        return {
            "primary_key": primary_key,
            "columns": [header for _, header in columns],
            "records": records,
            "errors": errors,
        }

    opportunities = [dict(o) for o in ledger.get("opportunities", [])]
    customers = []
    for customer in ledger.get("customers", []):
        row_data = dict(customer)
        row_data.update(_customer_aggregates(ledger, customer["cust_id"], year))
        row_data.setdefault("archive_updated_at", customer.get("archive_updated_at"))
        customers.append(row_data)

    tables = {
        "门店客户商机表": make_table("门店客户商机表", "opp_id", opportunities),
        "总部客户档案表": make_table("总部客户档案表", "cust_id", customers),
    }
    return {
        "schema_version": 1,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "source": "local_json",
        "tables": tables,
        "errors": [
            {"table": table_name, **error}
            for table_name, table in tables.items()
            for error in table["errors"]
        ],
    }


def cmd_export_sync_json(a, ledger, rules, today):
    payload = build_sync_record_sets(ledger, rules, today)
    if a.out:
        out_path = Path(a.out).expanduser().resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        return {
            "status": "exported",
            "out": str(out_path),
            "counts": {name: len(table["records"]) for name, table in payload["tables"].items()},
            "errors": payload["errors"],
            "_human": f"已导出飞书双表同步 JSON：{out_path}",
        }
    return payload


def cmd_sync_feishu(a, ledger, rules, today):
    """飞书多维表格双表同步（V1.2.9 新增）。

    调用 sync_feishu_bitable.run_sync()，默认 dry-run 只预览；
    带 --yes 实际写入飞书。依赖 lark-cli 已登录且对目标 Base 有权限。
    """
    try:
        from sync_feishu_bitable import run_sync
    except ImportError:
        return {"error": "dependency", "detail": "找不到 sync_feishu_bitable.py，请确认 scripts/ 目录完整"}

    # 透传参数
    base_token = getattr(a, "base_token", None) or None
    table_opp = getattr(a, "table_opp", None) or None
    table_cust = getattr(a, "table_cust", None) or None
    batch_size = getattr(a, "batch_size", 50) or 50
    dry_run = not getattr(a, "yes", False)

    result = run_sync(
        base_token=base_token,
        table_opp=table_opp,
        table_cust=table_cust,
        data=a.data,
        rules=a.rules,
        today=today,
        dry_run=dry_run,
        batch_size=batch_size,
        as_json=a.json,
    )

    if a.json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
        return {"_consumed": True}

    # 人类可读输出
    if not result.get("ok"):
        print(f"飞书同步失败（{result.get('stage', '?')}）: {result.get('error', result.get('detail', '?'))}")
        if result.get("hint"):
            print(f"提示: {result['hint']}")
        return {"_consumed": True}

    mode = "实际写入" if result["mode"] == "write" else "预览（dry-run）"
    print(f"飞书双表同步 · {mode}")
    print(f"Base: {result['base_token']}")
    for r in result["results"]:
        table = r.get("table", "?")
        if r.get("status") == "table_not_found":
            print(f"  [{table}] 表不存在: {r.get('error')}")
        elif r.get("status") == "error":
            print(f"  [{table}] 错误: {r.get('error')}")
        elif "total_local" in r:
            print(f"  [{table}] 本地 {r['total_local']} 条 | 远程已有 {r.get('existing_remote', '?')} | "
                  f"待新增 {r.get('to_create', 0)} | 待更新 {r.get('to_update', 0)}"
                  + (f" | 失败 {r.get('failed', 0)}" if r.get("failed") else ""))
            if r.get("missing_fields"):
                print(f"    缺少字段: {r['missing_fields']}")
            if r.get("errors"):
                for e in r["errors"]:
                    print(f"    错误: {e}")
    if result["mode"] == "dry-run":
        print("\n这是预览，未写入飞书。确认无误后加 --yes 执行实际写入。")
    return {"_consumed": True}


def cmd_export_xlsx(a, ledger, rules, today):
    if not a.out:
        return {"error": "missing_field", "detail": "export-xlsx 需要 --out 指定输出路径"}
    try:
        from openpyxl import Workbook
    except ImportError:
        return {"error": "dependency", "detail": "缺少 openpyxl，请先 pip install -r requirements.txt"}
    year = parse_date(today).year
    wb = Workbook()
    wb.remove(wb.active)

    exp = rules["export"]
    # 门店客户商机表
    cfg = exp["门店客户商机表"]
    ws = wb.create_sheet("门店客户商机表")
    ws.cell(row=1, column=1, value=cfg["title_row"])
    ws.cell(row=2, column=1, value=cfg["rule_row"])
    for ci, (_, header) in enumerate(cfg["columns"], start=1):
        ws.cell(row=cfg["header_row_index"], column=ci, value=header)
    r = cfg["header_row_index"] + 1
    for o in ledger["opportunities"]:
        for ci, (key, _) in enumerate(cfg["columns"], start=1):
            val = o.get(key)
            if key in ("quantity",):
                val = int(val) if isinstance(val, (int, float)) and val is not None else val
            ws.cell(row=r, column=ci, value=val)
        r += 1

    # 总部客户档案表
    cfg2 = exp["总部客户档案表"]
    ws2 = wb.create_sheet("总部客户档案表")
    ws2.cell(row=1, column=1, value=cfg2["title_row"])
    ws2.cell(row=2, column=1, value=cfg2["rule_row"])
    for ci, (_, header) in enumerate(cfg2["columns"], start=1):
        ws2.cell(row=cfg2["header_row_index"], column=ci, value=header)
    r = cfg2["header_row_index"] + 1
    for c in ledger["customers"]:
        agg = _customer_aggregates(ledger, c["cust_id"], year)
        row_data = dict(c)
        row_data.update(agg)
        row_data.setdefault("archive_updated_at", c.get("archive_updated_at"))
        for ci, (key, _) in enumerate(cfg2["columns"], start=1):
            ws2.cell(row=r, column=ci, value=row_data.get(key))
        r += 1

    out_path = Path(a.out).expanduser().resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {"_human": f"已导出双表：{out_path}（商机 {len(ledger['opportunities'])} 条 / 客户 {len(ledger['customers'])} 家）",
            "out": str(out_path), "opportunities": len(ledger["opportunities"]), "customers": len(ledger["customers"])}


def cmd_all(a, ledger, rules, today):
    return ledger


def cmd_quote(a, ledger, rules, today):
    qt = rules["quote_tool"]
    tool_path = (SKILL_DIR / qt["relpath"]).resolve()
    if not tool_path.exists():
        return {"error": "tool_missing", "detail": f"报价工具不存在：{tool_path}"}
    # 预算/盒
    budget = a.budget_per_box
    if budget is None and a.amount is not None and a.quantity:
        budget = round(a.amount / a.quantity, 2)
    if budget is None:
        return {"error": "missing_field",
                "detail": "组盒报价需要单盒含税运预算（--budget-per-box），或同时给 --amount 与 --quantity 由引擎推算"}
    if not a.customer:
        return {"error": "missing_field", "detail": "报价需要客户名称（--customer）"}
    if not a.order_quantity:
        return {"error": "missing_field", "detail": "报价需要订单盒数（--order-quantity）"}

    cmd = [qt["runtime"], str(tool_path),
           "--customer", a.customer,
           "--budget-per-box", str(budget),
           "--notes", a.notes or qt["defaults"]["notes"],
           "--order-quantity", str(a.order_quantity),
           "--delivery-area", a.delivery_area or qt["defaults"]["delivery_area"],
           "--invoice", a.invoice or qt["defaults"]["invoice"],
           "--box", a.box or qt["defaults"]["box"]]
    if a.freight_total is not None:
        cmd += ["--freight-total", str(a.freight_total)]
    if a.freight_per_box is not None:
        cmd += ["--freight-per-box", str(a.freight_per_box)]
    if a.image_dir:
        cmd += ["--image-dir", a.image_dir]
    if a.date:
        cmd += ["--date", a.date]
    if a.list_plans:
        cmd += ["--list-plans", str(a.list_plans)]
    elif a.plan_index:
        cmd += ["--alternatives", str(a.alternatives or 3), "--plan-index", str(a.plan_index)]
        if a.out:
            cmd += ["--output", a.out]

    if a.dry_run:
        return {"_human": "（dry-run）将执行：" + " ".join(cmd), "cmd": cmd, "budget_per_box": budget}

    try:
        proc = subprocess.run(cmd, capture_output=True, text=True)
    except FileNotFoundError as e:
        return {"error": "runtime_missing", "detail": f"无法运行 {qt['runtime']}：{e}"}
    result = {"cmd": cmd, "returncode": proc.returncode, "stdout": proc.stdout, "stderr": proc.stderr,
              "budget_per_box": budget}
    if proc.returncode != 0:
        result["error"] = "quote_failed"
        result["detail"] = "报价工具返回非零，详见 stderr"
        result["_human"] = f"报价失败（退出码 {proc.returncode}）：{(proc.stderr or proc.stdout or '').strip()[:500]}"
    else:
        result["_human"] = f"报价工具已执行。{('输出：' + a.out) if a.out and a.plan_index else '候选方案见 stdout。'}"
    return result


# ----------------------------- apply：契约 consumer -----------------------------

_NS_DEFAULTS = dict(
    opp_id=None, cust_id=None, customer_name=None, contact_phone=None, new_customer=False,
    name=None, short_name=None, contact=None, customer_type=None, industry=None,
    customer_level=None, region=None, address=None, title=None, other_contacts=None,
    preference=None, tags=None, customer_status=None, notes=None,
    source=None, need=None, product=None, quantity=None, amount=None,
    intent_level=None, stage=None, next_date=None, owner=None, store=None,
    content=None, feedback=None, next_action=None, followed_at=None, follow_content=None,
    deal_date=None, deal_amount=None, payment_status=None, lost_reason=None,
    customer_address=None, demand_category=None,
    payment_date=None, payment_amount=None, delivery_date=None,
    record=None, overdue=False, confirmed=False,
    expect_version=None, idempotency_key=None,
)


def namespace_from_record(record):
    """把一条抽取记录映射成 cmd_* 复用的参数命名空间。"""
    a = SimpleNamespace(**_NS_DEFAULTS)
    customer = record.get("customer") or {}
    opp = record.get("opportunity") or {}
    fu = record.get("follow_up") or {}
    deal = record.get("deal") or {}
    match = record.get("match") or {}

    a.store = record.get("store")
    a.owner = record.get("owner")
    a.source = record.get("source")
    a.expect_version = record.get("expect_version")
    a.opp_id = match.get("opp_id")
    a.cust_id = match.get("cust_id")
    a.customer_name = match.get("customer_name") or (customer.get("name") if record.get("intent") != "add_customer" else None)
    a.contact_phone = match.get("contact_phone") or customer.get("contact_phone")

    a.name = customer.get("name")
    a.short_name = customer.get("short_name")
    a.contact = customer.get("contact")
    a.customer_type = customer.get("customer_type")
    a.industry = customer.get("industry")
    a.customer_level = customer.get("customer_level")
    a.region = customer.get("region")
    a.address = customer.get("address")
    a.title = customer.get("title")
    a.other_contacts = customer.get("other_contacts")
    a.preference = customer.get("preference")
    a.tags = customer.get("tags")
    a.notes = customer.get("notes")

    a.need = opp.get("customer_need")
    a.product = opp.get("product")
    a.demand_category = opp.get("demand_category")
    a.customer_address = opp.get("customer_address") or customer.get("address")
    a.payment_date = opp.get("payment_date")
    a.payment_amount = opp.get("payment_amount")
    a.delivery_date = opp.get("delivery_date")
    a.quantity = opp.get("quantity")
    a.amount = opp.get("expected_amount")
    a.intent_level = opp.get("intent_level")
    a.stage = opp.get("stage")
    a.next_date = opp.get("next_follow_date")

    a.follow_content = fu.get("content")
    a.content = fu.get("content")
    a.feedback = fu.get("customer_feedback")
    a.next_action = fu.get("next_action")
    a.followed_at = fu.get("followed_at")

    a.deal_date = deal.get("deal_date")
    a.payment_status = deal.get("payment_status")
    a.lost_reason = deal.get("lost_reason")
    if deal.get("deal_amount") is not None:
        a.amount = deal.get("deal_amount")  # close 时以成交金额为准
    return a


def apply_record(ledger, rules, today, record, confirmed=False):
    intent = record.get("intent")
    a = namespace_from_record(record)
    a.confirmed = confirmed
    if intent == "add_customer":
        return cmd_add(a, ledger, rules, today)
    if intent == "update_opportunity":
        return cmd_update(a, ledger, rules, today)
    if intent == "log_followup":
        return cmd_log(a, ledger, rules, today)
    if intent == "close_deal":
        return cmd_close(a, ledger, rules, today)
    if intent == "update_payment":
        # V1.2：更新回款 = 进入已回款（需回款日期+回款金额）
        a.stage = "已回款"
        return cmd_close(a, ledger, rules, today)
    if intent == "update_delivery":
        # V1.2：仅更新交货日期
        if not a.delivery_date:
            return {"error": "missing_field", "detail": "update_delivery 必须提供交货日期（delivery_date）"}
        a.stage = None
        return cmd_update(a, ledger, rules, today)
    if intent == "set_next":
        # 仅改下次跟进日期
        a.stage = None
        return cmd_update(a, ledger, rules, today)
    if intent == "query":
        return cmd_query(a, ledger, rules, today)
    return {"error": "bad_intent", "detail": f"未知 intent：{intent}"}


def cmd_apply(a, ledger, rules, today):
    if not a.record:
        return {"error": "missing_field", "detail": "apply 需要 --record <record.json>"}
    p = Path(a.record)
    if not p.exists():
        return {"error": "not_found", "detail": f"找不到记录文件 {a.record}"}
    try:
        record = json.loads(p.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        return {"error": "bad_json", "detail": f"记录 JSON 解析失败：{e}"}
    return apply_record(ledger, rules, today, record, confirmed=a.confirmed)


# ----------------------------- 分派 -----------------------------

def _match(a):
    return {"opp_id": a.opp_id, "cust_id": a.cust_id,
            "customer_name": a.customer_name, "contact_phone": a.contact_phone}


WRITE_MODES = {"add", "update", "log", "close", "apply"}


def build_parser():
    p = argparse.ArgumentParser(description="门店团购客户与商机台账引擎（双表模型）")
    p.add_argument("--mode", "-m", required=True,
                   choices=["add", "update", "log", "close", "show", "due", "query",
                            "summary", "report", "review", "export-xlsx", "export-sync-json",
                            "sync-feishu",
                            "all", "quote", "apply"])
    p.add_argument("--data", default=str(DEFAULT_LEDGER), help="台账 JSON 路径")
    p.add_argument("--rules", default=str(RULES_PATH), help="规则文件路径")
    p.add_argument("--today", default=None, help="今天 YYYY-MM-DD（默认系统日期）")
    p.add_argument("--json", action="store_true", help="以 JSON 输出")
    p.add_argument("--confirmed", action="store_true",
                   help="店长已确认金额/数量/电话/日期；写入类操作含这些字段时必须带此标志，否则拒写")
    # 定位
    p.add_argument("--opp-id"); p.add_argument("--cust-id")
    p.add_argument("--customer-name"); p.add_argument("--contact-phone")
    p.add_argument("--new-customer", action="store_true", help="同名客户但确为新客户时强制新建 KH")
    # 客户字段
    p.add_argument("--name"); p.add_argument("--short-name"); p.add_argument("--contact")
    p.add_argument("--customer-type"); p.add_argument("--industry"); p.add_argument("--customer-level")
    p.add_argument("--region"); p.add_argument("--address"); p.add_argument("--title")
    p.add_argument("--other-contacts"); p.add_argument("--preference"); p.add_argument("--tags")
    p.add_argument("--customer-status"); p.add_argument("--notes")
    # 商机字段
    p.add_argument("--source"); p.add_argument("--need"); p.add_argument("--product")
    p.add_argument("--quantity", type=int); p.add_argument("--amount", type=float)
    p.add_argument("--intent-level"); p.add_argument("--stage"); p.add_argument("--next-date")
    p.add_argument("--owner"); p.add_argument("--store")
    # 跟进字段
    p.add_argument("--content"); p.add_argument("--feedback"); p.add_argument("--next-action")
    p.add_argument("--followed-at"); p.add_argument("--follow-content")
    # 成交/回款/交货字段（V1.2）
    p.add_argument("--deal-date"); p.add_argument("--deal-amount", type=float)
    p.add_argument("--payment-status"); p.add_argument("--lost-reason")
    p.add_argument("--customer-address"); p.add_argument("--demand-category")
    p.add_argument("--payment-date"); p.add_argument("--payment-amount", type=float)
    p.add_argument("--delivery-date")
    # 报表
    p.add_argument("--period", choices=["daily", "weekly", "monthly"], default="weekly")
    p.add_argument("--overdue", action="store_true")
    p.add_argument("--out")
    # 报价
    p.add_argument("--customer"); p.add_argument("--budget-per-box", type=float)
    p.add_argument("--order-quantity", type=int); p.add_argument("--delivery-area")
    p.add_argument("--invoice"); p.add_argument("--box"); p.add_argument("--list-plans", type=int)
    p.add_argument("--plan-index", type=int); p.add_argument("--alternatives", type=int)
    p.add_argument("--freight-total", type=float); p.add_argument("--freight-per-box", type=float)
    p.add_argument("--image-dir"); p.add_argument("--date"); p.add_argument("--dry-run", action="store_true")
    # apply（契约 consumer）
    p.add_argument("--record", help="apply 模式：已校验的抽取记录 JSON 文件")
    # 飞书同步（V1.2.9）
    p.add_argument("--base-token", default=None, help="sync-feishu：飞书 Base token")
    p.add_argument("--table-opp", default=None, help="sync-feishu：商机表名称")
    p.add_argument("--table-cust", default=None, help="sync-feishu：客户档案表名称")
    p.add_argument("--batch-size", type=int, default=50, help="sync-feishu：每批写入记录数（最大200）")
    p.add_argument("--yes", action="store_true", help="sync-feishu：实际写入飞书（默认 dry-run 预览）")
    # 可靠性（V1.2.2）
    p.add_argument("--expect-version", type=int, help="乐观锁：期望的商机版本号，不匹配则拒绝更新")
    p.add_argument("--idempotency-key", help="幂等键：相同key重复提交只执行一次")
    return p


def emit(result, as_json):
    if as_json:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    elif isinstance(result, dict) and "_human" in result:
        print(result["_human"])
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


def main(argv=None):
    args = build_parser().parse_args(argv)
    today = to_iso(args.today) or date.today().isoformat()
    rules = load_rules(args.rules)
    ledger = load_ledger(args.data)

    # TC-098：幂等键——重复提交直接返回首次结果，不再重复写入
    if args.mode in WRITE_MODES and args.idempotency_key:
        hit = ledger.get("idempotency", {}).get(args.idempotency_key)
        if hit:
            res = dict(hit.get("result") or {})
            res["status"] = res.get("status") or "idempotent_hit"
            res["detail"] = res.get("detail") or "相同幂等键的请求已处理过，本次未重复执行"
            print(json.dumps(res, ensure_ascii=False, indent=2, default=str))
            return 0

    mode_map = {
        "add": lambda: cmd_add(args, ledger, rules, today),
        "update": lambda: cmd_update(args, ledger, rules, today),
        "log": lambda: cmd_log(args, ledger, rules, today),
        "close": lambda: cmd_close(args, ledger, rules, today),
        "show": lambda: cmd_show(args, ledger, rules, today),
        "due": lambda: cmd_due(args, ledger, rules, today),
        "query": lambda: cmd_query(args, ledger, rules, today),
        "summary": lambda: cmd_summary(args, ledger, rules, today),
        "report": lambda: cmd_report(args, ledger, rules, today),
        "review": lambda: cmd_review(args, ledger, rules, today),
        "export-xlsx": lambda: cmd_export_xlsx(args, ledger, rules, today),
        "export-sync-json": lambda: cmd_export_sync_json(args, ledger, rules, today),
        "sync-feishu": lambda: cmd_sync_feishu(args, ledger, rules, today),
        "all": lambda: cmd_all(args, ledger, rules, today),
        "quote": lambda: cmd_quote(args, ledger, rules, today),
        "apply": lambda: cmd_apply(args, ledger, rules, today),
    }
    result = mode_map[args.mode]()
    if args.mode in WRITE_MODES and isinstance(result, dict) and result.get("_mutated") \
            and "error" not in result and result.get("status") != "needs_confirmation":
        if args.idempotency_key:  # TC-098：记录幂等结果与台账同事务写入
            ledger.setdefault("idempotency", {})[args.idempotency_key] = {
                "at": today, "mode": args.mode,
                "result": {k: v for k, v in result.items() if k != "_human"}}
        save_ledger(args.data, ledger)
    # sync-feishu 等模式自己控制输出，标记 _consumed 后不再 emit
    if not (isinstance(result, dict) and result.get("_consumed")):
        emit(result, args.json)
    return 0


if __name__ == "__main__":
    sys.exit(main())
